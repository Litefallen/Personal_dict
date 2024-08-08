from openpyxl import Workbook, load_workbook
import os

def open_xlsx_file():
    print('Please, provide the full path to the xlsx file,')
    the_path = input('Or press "enter" to use default app folder: ').replace('\\','/') # replace windows backslashes in path
    if the_path:
        os.chdir(the_path)
    try:  # check is xlsx file exist
        wb = load_workbook('dictionary.xlsx')
    except FileNotFoundError:
        print('No file found')
        wb = Workbook()
        wb.save('dictionary.xlsx')  # create one if it doesnt exist
    try:
        # check if workbook exists inside the xlsx file, if no - create one
        ws = wb['Dictionary']
    except KeyError:
        print('No "Dictionary" sheet found')
        ws = wb.active
        ws.title = 'Dictionary'
        ws['A1'] = 'word_number'
        ws['B1'] = 'english_word'
        ws['C1'] = 'word_meaning'
        ws['D1'] = 'russian_translation'
        wb.save('dictionary.xlsx')
    dictionary_workbook = wb
    print(f'The xlsx  file tith all words could be found in "{os.getcwd()}" folder.')
    return dictionary_workbook


def save_word(word_m_t: tuple,workbook):
    wb = workbook
    ws = wb['Dictionary']
    max_row = ws.max_row
    # save initial word, meaning and translation to cells
    ws[f'A{max_row+1}'] = max_row
    ws[f'B{max_row+1}'] = word_m_t[0]
    ws[f'C{max_row+1}'] = word_m_t[1]
    ws[f'D{max_row+1}'] = word_m_t[2]
    wb.save('dictionary.xlsx')


def get_all_words(workbook):
    wb = workbook
    ws = wb['Dictionary']
    all_words = [i for i in ws.values][1:]
    # add only one meaning of the word for not cluttering terminal screen sake 
    all_words = [[i[0], i[1], sorted(i[2].split(';'), key=len)[
        0].strip(',;. '), i[3]] for i in all_words]
    return all_words